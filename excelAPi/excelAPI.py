

import re
from openpyxl.utils import  coordinate_to_tuple


def Get_numberic_axis(coord):
   return_list = []
   coord = coord.replace("$","")
   lst = coord.split(":")
   for each in lst:
      return_list.append(coordinate_to_tuple(each))
   return return_list


def Load_definedname(workbook):
   return_dict = dict()
   namelist = workbook.defined_names.definedName
   
   for each in namelist:
      temporary_list = list()
      name = each.name
      for sheet,coord in each.destinations:
         temporary_list.append(workbook[sheet])
         temporary_list.append(coord)
      return_dict[each.name] = temporary_list
   return return_dict

def Get_cell_value(worksheet,string_pos=None,column=None,row=None):
   if string_pos is not None:
      return worksheet[string_pos].value
   else:
      return worksheet.cell(row,column).value

def Set_cell_value(worksheet,value,string_pos=None,column=None,row=None):
   if string_pos is not None:
      worksheet[string_pos] = value
   else:
      worksheet.cell(row,column).value = value










   
   
      

 


   


# print(name)
# lst = list(wb.get_named_ranges())
# for each in lst:
#    print(lst)
# print(list(wb.defined_names))


# for j in range(sheet.min_column,sheet.max_column+1):
#    for j in range(sheet.min_row,sheet.max_row+1):

   

 





